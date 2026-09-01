"""
Bibliothèque de rendu des classeurs de migration.

Ce module ne connaît aucun métier : il rend une spécification de colonnes en
classeur Excel prêt à faire remplir. La spécification (quelles feuilles,
quelles colonnes, quelles listes) vit dans le `build_template.py` du projet,
qui importe ce module.

À copier dans le dossier de migration du projet, à côté du `build_template.py`,
pour que le dossier reste rejouable sans le skill.

Format rendu par `build_sheet` :
    ligne 1  bandeau de groupe (fusionné, fond de couleur, texte blanc)
    ligne 2  nom de la colonne, suffixé « * » si obligatoire
    ligne 3  aide de saisie (format attendu, exemple)
    ligne 4+ zone de saisie, avec listes déroulantes et formats de date

Dépendance : openpyxl.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DATA_ROWS = 250

# ---------------------------------------------------------------------------
# Palette
# ---------------------------------------------------------------------------

# Couleur par nom de groupe. Les groupes absents de ce dict reçoivent une
# couleur de la palette de repli, dans leur ordre d'apparition.
GROUP_COLORS = {
    "Identité": "1F3A2E",
    "Contact principal": "2F5D50",
    "Adresse": "2F5D50",
    "Origine": "4A7C6F",
    "Pipe commercial": "8C6A3F",
    "Contrat": "8C6A3F",
    "Équipe": "3B5B7A",
    "Organisation": "3B5B7A",
    "Suivi": "6B7280",
    "Divers": "6B7280",
}

FALLBACK_COLORS = [
    "1F3A2E", "3B5B7A", "8C6A3F", "6B4E71", "A05B3C", "4A6B8A", "4A7C6F",
    "6B7280",
]

HEADER_FILL = "EDEFF1"
HELP_FILL = "FAFAFA"
HELP_COLOR = "6B7280"
LABEL_COLOR = "1F2933"
REQUIRED_COLOR = "9A3412"
TITLE_COLOR = "1F3A2E"
THIN = Side(style="thin", color="D4D8DD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DATE_HELP = "AAAA-MM-JJ"
DATE_FMT = "yyyy-mm-dd"

# Valeur de `dv` qui produit une liste Oui/Non sans passer par la feuille Listes.
OUINON = "__ouinon"

COMMON_RULES = [
    "Ligne 1 = le bloc, ligne 2 = le nom de la colonne, ligne 3 = l'aide. "
    "Commencer à saisir en ligne 4, une ligne par élément.",
    "Les colonnes suivies d'une astérisque (*) sont obligatoires. Le reste peut "
    "rester vide et sera complété plus tard dans l'outil.",
    "Ne pas renommer, déplacer ni supprimer de colonne : l'import se cale dessus.",
    "Les cellules avec une liste déroulante n'acceptent que les valeurs proposées. "
    "S'il manque une valeur, laisser vide et le signaler plutôt que d'écrire à côté.",
    "Dates au format AAAA-MM-JJ (ex. 2026-03-14). Montants en nombres entiers, "
    "sans symbole ni espace (ex. 3500).",
]


# ---------------------------------------------------------------------------
# Spécification
# ---------------------------------------------------------------------------


def field(label, db, help_text="", width=22, dv=None, required=False, fmt=None):
    """Décrit une colonne.

    label     Libellé humain, lu par l'import. Unique dans la feuille.
    db        Cible technique (« table.colonne »). Sert de mapping, jamais rendu.
    help_text Format attendu et exemple. Court : la cellule fait 3 lignes.
    width     Largeur de colonne, en caractères.
    dv        Clé d'une liste passée à `add_lists_sheet`, ou OUINON.
    required  Ajoute « * » au libellé et le colore.
    fmt       Format de nombre openpyxl appliqué à la zone de saisie (DATE_FMT).
    """
    return {
        "label": label,
        "db": db,
        "help": help_text,
        "width": width,
        "dv": dv,
        "required": required,
        "fmt": fmt,
    }


# ---------------------------------------------------------------------------
# Rendu
# ---------------------------------------------------------------------------


def new_workbook():
    """Classeur vide, sans la feuille par défaut."""
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def add_lists_sheet(wb, lists):
    """Pose la feuille technique des listes déroulantes, et la masque.

    lists   dict {clé: [valeurs]}, dans l'ordre voulu.
    Retour  dict {clé: référence de plage}, à passer à `build_sheet`.
    """
    ws = wb.create_sheet("Listes")
    ranges = {}
    for idx, (key, values) in enumerate(lists.items(), start=1):
        col = get_column_letter(idx)
        ws[f"{col}1"] = key
        ws[f"{col}1"].font = Font(bold=True)
        for row, value in enumerate(values, start=2):
            ws[f"{col}{row}"] = value
        ranges[key] = f"Listes!${col}$2:${col}${len(values) + 1}"
        ws.column_dimensions[col].width = 26
    ws.sheet_state = "hidden"
    return ranges


def _group_color(name, seen):
    if name in GROUP_COLORS:
        return GROUP_COLORS[name]
    if name not in seen:
        seen[name] = FALLBACK_COLORS[len(seen) % len(FALLBACK_COLORS)]
    return seen[name]


def build_sheet(wb, title, groups, ranges, data_rows=DATA_ROWS):
    """Rend une feuille de saisie.

    groups  liste de (nom du groupe, [field(), ...]), dans l'ordre des colonnes.
    ranges  sortie de `add_lists_sheet`.
    """
    ws = wb.create_sheet(title)
    seen_colors = {}
    validations = {}
    col = 1

    for group_name, fields in groups:
        color = _group_color(group_name, seen_colors)
        start = col

        for f in fields:
            letter = get_column_letter(col)

            banner = ws.cell(row=1, column=col)
            banner.fill = PatternFill("solid", fgColor=color)
            banner.border = BORDER

            head = ws.cell(row=2, column=col)
            head.value = f["label"] + (" *" if f["required"] else "")
            head.font = Font(bold=True, size=10,
                             color=REQUIRED_COLOR if f["required"] else LABEL_COLOR)
            head.fill = PatternFill("solid", fgColor=HEADER_FILL)
            head.alignment = Alignment(vertical="center", wrap_text=True)
            head.border = BORDER

            hint = ws.cell(row=3, column=col)
            hint.value = f["help"]
            hint.font = Font(italic=True, size=8, color=HELP_COLOR)
            hint.fill = PatternFill("solid", fgColor=HELP_FILL)
            hint.alignment = Alignment(vertical="top", wrap_text=True)
            hint.border = BORDER

            ws.column_dimensions[letter].width = f["width"]

            if f["fmt"]:
                for row in range(4, 4 + data_rows):
                    ws.cell(row=row, column=col).number_format = f["fmt"]

            if f["dv"]:
                formula = '"Oui,Non"' if f["dv"] == OUINON else ranges[f["dv"]]
                validations.setdefault(formula, []).append(
                    f"{letter}4:{letter}{3 + data_rows}"
                )

            col += 1

        cell = ws.cell(row=1, column=start)
        cell.value = group_name
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col - 1 > start:
            ws.merge_cells(start_row=1, start_column=start, end_row=1,
                           end_column=col - 1)

    # Une DataValidation par liste, plutôt qu'une par colonne : Excel râle
    # au-delà de quelques dizaines de validations sur une même feuille.
    for formula, cells in validations.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=True,
                            showDropDown=False)  # False => la flèche s'affiche
        dv.error = "Choisir une valeur dans la liste déroulante."
        dv.errorTitle = "Valeur hors liste"
        ws.add_data_validation(dv)
        for ref in cells:
            dv.add(ref)

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 34
    ws.freeze_panes = "B4"
    return ws


def build_readme(wb, title, intro, steps, rules):
    """Pose l'onglet LISEZ-MOI en première position.

    intro  paragraphes : ce que contient le fichier, quand le remplir.
    steps  marche à suivre, numérotée.
    rules  contraintes ; passer COMMON_RULES + les règles propres au fichier.
    """
    ws = wb.create_sheet("LISEZ-MOI", 0)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 110
    ws.sheet_view.showGridLines = False

    row = 2
    ws[f"B{row}"] = title
    ws[f"B{row}"].font = Font(bold=True, size=16, color=TITLE_COLOR)
    row += 2

    for line in intro:
        ws[f"B{row}"] = line
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 30
        row += 1

    def section(heading, lines):
        nonlocal row
        row += 1
        ws[f"B{row}"] = heading
        ws[f"B{row}"].font = Font(bold=True, size=12, color=TITLE_COLOR)
        row += 1
        for line in lines:
            ws[f"B{row}"] = line
            ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = 30
            row += 1

    section("Comment remplir",
            [f"{i}.  {s}" for i, s in enumerate(steps, start=1)])
    section("Règles à respecter", [f"•  {r}" for r in rules])
    return ws
